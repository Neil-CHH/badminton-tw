# -*- coding: utf-8 -*-
"""tsbadminton 賽程表/秩序冊 xlsx 的解析工具。

這些 xlsx 是「空間排版的籤表樹」,無法還原成逐場比分,但兩件事很有價值:

1. **參賽名冊** — 每個籤位都是「序號｜單位｜姓名」的相鄰儲存格,是真文字。
   成績只有 JPG 圖片,視覺解析容易讀錯字;有名冊就能把「自由辨識」降成
   「從已知名單挑一個」,精確度差很多(見 tsba_reconcile.py)。
2. **比賽日期** — 日賽程分頁有真正的比賽日期,列表頁只有公告有效期(非賽期)。

需要 openpyxl。
"""
import datetime
import re

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

# 組別標題。兩種寫法都要認:會長盃用短寫法「U9男單」,清晨盃用長寫法
# 「30歲男子甲組單打」「專業組29歲以下男單」。
EVENT_PAT = re.compile(
    r"(男單|女單|男雙|女雙|混雙|男團|女團|團體|"
    r"(?:男子|女子|混合|父子|母子|祖孫|親子)[^,，。]{0,6}(?:單打|雙打|團體))")
_SEED = re.compile(r"\s*[\[［]\s*\d+\s*[\]］]\s*$")   # 姓名後的種子序號
_BYE = re.compile(r"^(bye|輪空)", re.I)
_TIME = re.compile(r"^\d{1,2}:\d{2}")
_NOISE = re.compile(r"^[\d\s:：/\-]+$")
# 姓名不會有標點;單位可以(跨社雙打寫成「A社,B社」),所以兩者用不同規則。
_PUNCT = re.compile(r"[,，。;；、!！?？…]")
# 「共 80 組,80 場,取 4 名」被 PDF 切開後的碎片,例如「組,80」「場,取」
_COUNT_FRAG = re.compile(r"^(共|取|組|場|名|隊)\s*[，,、]|[，,]\s*\d+\s*[組場名]|^\d+\s*[組場名]$")
# 標題常帶流水號前綴與人數說明:「27、 30歲男子甲組單打」「4、專業組29歲以下男單 ：118人」
_TITLE_PREFIX = re.compile(r"^\s*\d+\s*[、.．]\s*")
_TITLE_TAIL = re.compile(r"\s*[：:(（].*$")
_BLOCK_SUFFIX = re.compile(r"\s*\d+\s*[-–]\s*\d+\s*$")   # 「U9男單 9-1」的區塊編號
# 宣告人數:「共 11 人」「：118人，117場」「21組」
_DECLARED = re.compile(r"(?:共\s*)?(\d+)\s*([人組隊])")
# 彙總類分頁,不當作宣告人數的來源
_SKIP_SHEET = re.compile(r"統計|公告|成績|總賽程|名單")
# 機構/球隊字樣。直式排版靠「上下相鄰」判斷,單位若不在詞彙裡會被誤當成姓名,
# 用這個把學校/球隊名擋掉(實測「林口國小｜新北江翠國小」被收成一組)。
_UNIT_WORD = re.compile(
    r"國小|國中|高中|高工|高商|國民|實中|中學|大學|學校|學院|球隊|羽球|俱樂部|"
    r"協會|中心|訓練|體育|聯隊|代表隊|分齡|團體")
# 團體隊伍名單的標籤(邀請組/VIP團體組用這種排版)
_TEAM_NAME = re.compile(r"^(隊名|隊伍|團隊)\s*[：:]")
_TEAM_MEMBER = re.compile(r"^(隊員|領隊|教練|隊長|選手)\s*[：:]")
# 表頭字樣,避免把「組別｜冠軍」這種標題列當成「單位｜姓名」
_HEADER = {"組別", "組  別", "編號", "姓名", "單位", "名次", "日期", "時間", "場次",
           "備註", "隊名", "選手", "冠軍", "亞軍", "季軍", "殿軍", "第五名", "場地",
           "序號", "隊伍", "學校", "球隊", "組別名稱", "領隊", "教練", "隊長", "隊員",
           "st.", "round", "winner", "final", "semifinals", "quarterfinals"}


def _cells(row):
    return [("" if v is None else str(v).strip()) for v in row]


def _is_header_word(s):
    return s.replace(" ", "").replace("　", "").lower() in _HEADER


def _is_name(s):
    """姓名:2-12 字、無數字、非時間、非 Bye(英文名如 Weijun Peng 也要收)。"""
    if not (2 <= len(s) <= 12) or _BYE.match(s) or _TIME.match(s):
        return False
    if _is_header_word(s) or _TEAM_NAME.match(s) or _TEAM_MEMBER.match(s):
        return False
    if _UNIT_WORD.search(s) or _PUNCT.search(s):   # 學校/球隊名稱、說明文字都不是姓名
        return False
    return not any(ch.isdigit() for ch in s) and ":" not in s and "：" not in s


def _is_unit(s):
    if _is_header_word(s) or EVENT_PAT.search(s) or _COUNT_FRAG.search(s):
        return False
    return 1 < len(s) <= 16 and not _BYE.match(s) and not _NOISE.match(s)


def clean_title(s):
    """把標題正規化成組別名:去流水號前綴、去「：118人…」說明、去區塊編號。"""
    s = _TITLE_PREFIX.sub("", str(s or "").strip())
    s = _TITLE_TAIL.sub("", s)
    s = _BLOCK_SUFFIX.sub("", s)
    s = re.sub(r"\s+", "", s).strip()
    s = re.sub(r"^[-–—、.．,，:：]+", "", s)   # PDF 切詞後常在開頭留下標點
    return re.sub(r"(參賽)?名單$", "", s).strip()


def _title_of(cell):
    """這格是不是組別標題?是就回傳正規化後的組別名。

    說明常和標題擠在同一格(「專業組29歲以下男單 ：108 人,107 場…」共 45 字),
    所以長度門檻要在「剝掉說明之後」才判斷。
    """
    raw = str(cell or "").strip()
    if not raw or len(raw) > 80 or not EVENT_PAT.search(raw):
        return None
    name = clean_title(raw)
    return name if 2 <= len(name) <= 30 else None


def _xlsx_sheets(path, sheets=None):
    if openpyxl is None:
        raise RuntimeError("需要 openpyxl:pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    for sname in (sheets or wb.sheetnames):
        if sname in wb.sheetnames:
            out[sname] = [_cells(r) for r in wb[sname].iter_rows(values_only=True)]
    wb.close()
    return out


def _pdf_sheets(path, y_tol=4, x_gap=3.5):
    """PDF 籤表 → 與 xlsx 相同的「分頁 → 列 → 格」結構。

    官方的賽程表有 xlsx 也有 PDF 兩種發布方式,版面完全一樣(團體名單同樣是
    `隊名：`/`隊員：`),所以轉成同一個結構後就能共用四種排版的解析邏輯。
    依 y 座標分列;同一列中橫向間距小於 x_gap 的詞併成一格,才不會把
    「25歲混合乙組雙打」拆成「25」與「歲混合乙組雙打」兩格。
    """
    try:
        import fitz  # noqa: PLC0415
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("需要 pymupdf:pip install pymupdf") from e
    doc = fitz.open(path)
    out = {}
    for pno in range(doc.page_count):
        buckets = {}
        for x0, y0, x1, _y1, word, *_ in doc[pno].get_text("words"):
            w = str(word).strip()
            if w:
                buckets.setdefault(round(y0 / y_tol), []).append((x0, x1, w))
        rows = []
        for _, ws in sorted(buckets.items()):
            cells, prev_x1 = [], None
            for x0, x1, w in sorted(ws):
                if prev_x1 is not None and x0 - prev_x1 <= x_gap:
                    cells[-1] += w
                else:
                    cells.append(w)
                prev_x1 = x1
            rows.append(cells)
        if rows:
            out[f"p{pno + 1}"] = rows
    doc.close()
    return out


def extract_roster(path, sheets=None):
    """回傳 {組別: [(單位, 姓名)]} —— 賽程表籤表裡的參賽名冊(吃 .xlsx 或 .pdf)。

    要處理四種排版(實測 2022~2026 的會長盃與清晨盃):
    A. 籤位列 `序號｜單位｜姓名`(雙打的搭檔可能同列並排,也可能在上一列)
    B. 沒填單位的籤位 `序號｜(空)｜姓名`(清晨盃的個人參賽者)
    C. 團體隊伍名單 `隊名：｜X隊` + `隊員：｜甲｜乙…`(邀請組/VIP團體組)
    D. 直式籤位:單位／姓名1／姓名2 疊在同一欄的連續列(2022 的雙打分頁)
    """
    is_pdf = str(path).lower().endswith(".pdf")
    sheet_rows = _pdf_sheets(path) if is_pdf else _xlsx_sheets(path, sheets)
    # PDF 的一個組別籤表會跨好幾頁,只有第一頁有標題,所以標題要延續到後續頁;
    # xlsx 的分頁彼此獨立,不能延續。
    carry_title = is_pdf

    # 單位詞彙:由「籤位序號｜單位｜姓名」建立,但要**整本活頁簿一起建**——
    # 2022 的雙打分頁是直式排版,自己那張表建不出詞彙,得借用單打分頁的。
    # 不能只靠相鄰位置判斷,否則並排的「姓名1｜姓名2」會被當成「單位｜姓名」
    # (實測 2023 誤判 656 筆)。
    units = set()
    for rows in sheet_rows.values():
        for cells in rows:
            for i in range(len(cells) - 2):
                seq, u, n = cells[i], cells[i + 1], cells[i + 2]
                if seq.isdigit() and _is_unit(u) and _is_name(_SEED.sub("", n).strip()):
                    units.add(u)

    roster = {}

    def add(group, unit, name):
        nm = _SEED.sub("", name).strip()
        if not _is_name(nm):
            return False
        roster.setdefault(group, [])
        if (unit, nm) not in roster[group]:
            roster[group].append((unit, nm))
        return True

    last_title = None
    for sname, rows in sheet_rows.items():
        # 先算出每一列所屬的組別(標題出現後一路沿用到下一個標題)
        titles, cur = [], (last_title if carry_title and last_title else sname)
        for cells in rows:
            for c in cells:
                t = _title_of(c)
                if t:
                    cur = t
                    break
            titles.append(cur)
        if carry_title and titles:
            last_title = titles[-1]

        current = sname
        team_unit = None      # 排版 C 目前所在的隊伍
        in_members = False    # 隊員名單可跨列延續
        for ri, cells in enumerate(rows):
            current = titles[ri]
            if ri and titles[ri] != titles[ri - 1]:
                team_unit, in_members = None, False

            head = cells[0] if cells else ""
            # --- 排版 C:團體隊伍名單 ---
            # PDF 版的標籤與值會擠在同一格(「隊名：苗栗縣校長隊」),要先切開;
            # xlsx 版則是分成兩格,所以兩種都要接。
            def _after_label(text):
                v = re.sub(r"^[^：:]*[：:]\s*", "", text).strip()
                return [v] if v else []

            if _TEAM_NAME.match(head):
                vals = _after_label(head) + [c for c in cells[1:] if c]
                team_unit = vals[0] if vals else None
                in_members = False
                continue
            if _TEAM_MEMBER.match(head):
                in_members = True
                for c in _after_label(head) + [c for c in cells[1:] if c]:
                    if team_unit:
                        add(current, team_unit, c)
                continue
            if in_members and team_unit and not head:
                # 隊員續列(第一格空白,後面接著名字)
                got = [c for c in cells[1:] if c]
                if got and all(_is_name(_SEED.sub("", c).strip()) for c in got):
                    for c in got:
                        add(current, team_unit, c)
                    continue
                in_members = False

            # --- 排版 A/B:籤位列 ---
            pos = next((i for i, u in enumerate(cells) if u and u in units), None)
            if pos is not None:
                unit = cells[pos]
                for nxt in cells[pos + 1:pos + 5]:
                    if not nxt:
                        continue
                    nm = _SEED.sub("", nxt).strip()
                    if not _is_name(nm) or nm in units:
                        break
                    add(current, unit, nm)
                continue
            # 沒填單位的籤位:序號 | (空) | 姓名
            for i in range(len(cells) - 2):
                if cells[i].isdigit() and not cells[i + 1] and cells[i + 2]:
                    add(current, "", cells[i + 2])
                    break

        # --- 排版 D:直式籤位(單位／姓名／姓名 疊在同一欄的連續列)---
        width = max((len(r) for r in rows), default=0)
        for col in range(width):
            ri = 0
            while ri < len(rows):
                cell = rows[ri][col] if col < len(rows[ri]) else ""
                if not cell or cell not in units:
                    ri += 1
                    continue
                k, taken = ri + 1, 0
                while k < len(rows) and taken < 2:
                    nxt = rows[k][col] if col < len(rows[k]) else ""
                    if not nxt:
                        break
                    nm = _SEED.sub("", nxt).strip()
                    if not _is_name(nm) or nm in units:
                        break
                    add(titles[k], cell, nm)
                    taken += 1
                    k += 1
                ri = k if taken else ri + 1

    return roster


def declared_counts(path, sheets=None):
    """回傳 {組別: (數量, '人'|'組')} —— 標題自己宣告的參賽規模,用來驗證抽取完整度。

    例:「4、專業組29歲以下男單 ：118人,117場」「2、U9男雙:21組」。
    """
    sheet_rows = (_pdf_sheets(path) if str(path).lower().endswith(".pdf")
                  else _xlsx_sheets(path, sheets))
    out = {}
    for sname, rows in sheet_rows.items():
        # 統計表/公告那類彙總分頁也有同名標題,但旁邊的數字是場數之類的別的東西
        # (實測「專業組29歲以下男單」在統計表旁邊是「421 組」,籤表標題才是「108 人」),
        # 會蓋掉正確答案,所以跳過。
        if _SKIP_SHEET.search(sname):
            continue
        for cells in rows:
            for i, c in enumerate(cells):
                title = _title_of(c)
                if not title or title in out:
                    continue
                # 人數可能和標題擠在同一格,也可能在同列較右邊的欄位(實測在第 6 欄)
                for txt in [c] + cells[i + 1:]:
                    m = _DECLARED.search(str(txt))
                    if m:
                        out[title] = (int(m.group(1)), m.group(2))
                        break
    return out


def extract_dates(path, year=None):
    """回傳 (dateStart, dateEnd)。掃所有分頁的日期型儲存格,取最小/最大。

    year 給定時只採該年度的日期,濾掉版本日期之類的雜訊。
    """
    if openpyxl is None:
        raise RuntimeError("需要 openpyxl:pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    found = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                d = None
                if isinstance(v, datetime.datetime):
                    d = v.date()
                elif isinstance(v, datetime.date):
                    d = v
                elif isinstance(v, str):
                    m = re.match(r"^(20\d\d)[-/](\d{1,2})[-/](\d{1,2})", v.strip())
                    if m:
                        try:
                            d = datetime.date(*(int(x) for x in m.groups()))
                        except ValueError:
                            d = None
                if d and (year is None or d.year == int(year)):
                    found.add(d)
    wb.close()
    if not found:
        return "", ""
    return min(found).isoformat(), max(found).isoformat()
